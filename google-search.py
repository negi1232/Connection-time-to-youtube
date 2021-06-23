# coding:utf-8
import time
from selenium import webdriver
import chromedriver_binary
import requests
from bs4 import BeautifulSoup
import time
import re
import lxml.html
import time
import datetime
import openpyxl
from selenium.webdriver.chrome.options import Options

if __name__ == "__main__":
    options = Options()
    #options.binary_location = 'C:\chromedriver_win32\chromedriver.exe'
    #options.add_argument('--headless')
    workbook = openpyxl.load_workbook('youtube.xlsx')

    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait = 120  
    sheet = workbook["Sheet1"]
    
    print(type(sheet))
    nowline=sheet['C1'].value
    while(True):
        print(nowline)
        now = datetime.datetime.today()
        start = time.time()
        driver.get('https://www.youtube.com/')
        elapsed_time = time.time() - start
        #print(now,end="")
        print(elapsed_time)
        
        workbook.save("youtube.xlsx")
        sheet.cell(row=nowline+1, column=1, value=str(now.hour)+":"+str(now.minute))
        sheet.cell(row=nowline+1, column=2, value=elapsed_time)
        nowline+=1   
        sheet.cell(row=1, column=3, value=nowline)
        
        workbook.save("youtube.xlsx")
        time.sleep(5)
